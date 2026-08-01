"""Export writers. AUDIT §8: answer_to_docx had no markdown parsing, no tables,
no hyperlinks and no bibliography despite citations being mandated."""

from __future__ import annotations

import io
import zipfile

from httpx import AsyncClient

from kimi.exports.markdown import BlockKind, parse_inline, parse_markdown
from kimi.exports.writers import (
    answer_to_docx,
    conversation_to_json,
    conversation_to_markdown,
    rows_to_xlsx,
    safe_stem,
    sources_to_csv,
    timestamped,
)

SOURCES = [
    {
        "index": 1,
        "title": "Libya oil output rises",
        "publisher": "reuters.com",
        "url": "https://reuters.com/world/africa/story-2026",
        "published_at": "2026-07-31T09:00:00+00:00",
        "status_label": "Full article read",
    }
]


def docx_text(data: bytes) -> str:
    """The raw document.xml, which is where structure is visible."""
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        return archive.read("word/document.xml").decode("utf-8")


# ---- markdown parsing -----------------------------------------------------


def test_headings_lists_and_tables_are_recognised() -> None:
    blocks = parse_markdown(
        "# Title\n\nIntro text.\n\n- one\n- two\n\n1. first\n\n| A | B |\n|---|---|\n| 1 | 2 |\n"
    )
    kinds = [b.kind for b in blocks]
    assert BlockKind.HEADING in kinds
    assert BlockKind.BULLET in kinds
    assert BlockKind.NUMBERED in kinds
    assert BlockKind.TABLE in kinds

    table = next(b for b in blocks if b.kind is BlockKind.TABLE)
    assert len(table.rows) == 2
    assert table.rows[0][0][0].text == "A"


def test_inline_styles_and_links() -> None:
    spans = parse_inline("plain **bold** and [link](https://example.com/x) and `code`")
    assert any(s.bold for s in spans)
    assert any(s.code for s in spans)
    link = next(s for s in spans if s.href)
    assert link.href == "https://example.com/x"
    assert link.text == "link"


def test_fenced_code_is_kept_verbatim() -> None:
    blocks = parse_markdown("```python\nx = 1\n# not a heading\n```")
    code = next(b for b in blocks if b.kind is BlockKind.CODE)
    assert code.language == "python"
    assert "# not a heading" in code.text


# ---- docx -----------------------------------------------------------------


def test_docx_renders_real_headings_not_literal_hashes() -> None:
    data = answer_to_docx(title="Report", body_markdown="## Findings\n\nBody text.")
    xml = docx_text(data)
    assert "Findings" in xml
    # The prototype shipped "## Findings" as literal text.
    assert "## Findings" not in xml
    assert "Heading" in xml


def test_docx_renders_a_real_table() -> None:
    data = answer_to_docx(
        title="R", body_markdown="| Field | Output |\n|---|---|\n| Sharara | 300000 |"
    )
    xml = docx_text(data)
    assert "<w:tbl>" in xml  # a genuine Word table
    assert "Sharara" in xml
    assert "|" not in xml.split("<w:body>")[1][:4000]  # not pipe soup


def test_docx_includes_a_clickable_hyperlink() -> None:
    data = answer_to_docx(title="R", body_markdown="See [the story](https://example.com/story).")
    xml = docx_text(data)
    assert "<w:hyperlink" in xml
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        rels = archive.read("word/_rels/document.xml.rels").decode()
    assert "https://example.com/story" in rels


def test_docx_writes_a_sources_bibliography() -> None:
    """The mandated [n] markers had no bibliography in the prototype."""
    data = answer_to_docx(title="R", body_markdown="Output rose [1].", sources=SOURCES)
    xml = docx_text(data)
    assert "Sources" in xml
    assert "Libya oil output rises" in xml
    assert "reuters.com" in xml
    assert "Full article read" in xml


def test_docx_marks_arabic_paragraphs_rtl() -> None:
    data = answer_to_docx(title="تقرير", body_markdown="ارتفع إنتاج النفط في ليبيا.")
    xml = docx_text(data)
    assert "w:bidi" in xml
    assert "ليبيا" in xml


def test_docx_bullets_use_the_list_style() -> None:
    data = answer_to_docx(title="R", body_markdown="- alpha\n- beta")
    xml = docx_text(data)
    assert "ListBullet" in xml.replace(" ", "")
    assert "- alpha" not in xml


# ---- other formats --------------------------------------------------------


def test_markdown_export_includes_citations() -> None:
    text = conversation_to_markdown(
        title="Libya",
        messages=[{"role": "assistant", "content": "Output rose [1].", "citations": SOURCES}],
    )
    assert "# Libya" in text
    assert "Output rose [1]." in text
    assert "reuters.com" in text


def test_json_export_is_timezone_aware_and_round_trippable() -> None:
    import json

    raw = conversation_to_json(
        title="T", messages=[{"role": "user", "content": "hi", "citations": None}]
    )
    payload = json.loads(raw)
    assert payload["schema"] == "kimi.conversation/1"
    # datetime.utcnow() was deprecated and naive; this must carry an offset.
    assert payload["exported_at"].endswith("+00:00")
    assert payload["messages"][0]["content"] == "hi"


def test_csv_export_lists_provenance_columns() -> None:
    text = sources_to_csv(SOURCES)
    assert "publisher" in text.splitlines()[0]
    assert "reuters.com" in text


def test_xlsx_export_is_a_real_workbook() -> None:
    from openpyxl import load_workbook

    data = rows_to_xlsx({"Sources": [["a", "b"], [1, 2]]})
    book = load_workbook(io.BytesIO(data))
    assert book.sheetnames == ["Sources"]
    assert book["Sources"].cell(1, 1).value == "a"


def test_filenames_are_safe_and_timestamped() -> None:
    assert safe_stem("Libya: oil/output report!") == "libya-oiloutput-report"
    assert safe_stem("") == "kimi-export"
    name = timestamped("report", "docx")
    assert name.startswith("report-") and name.endswith(".docx")


# ---- endpoints ------------------------------------------------------------


async def test_export_conversation_as_docx(client: AsyncClient, conversation_id: str) -> None:
    await client.post(
        "/api/chat/stream", json={"conversation_id": conversation_id, "content": "hello"}
    )
    resp = await client.get(
        f"/api/exports/conversation/{conversation_id}", params={"format": "docx"}
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument")
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.content[:2] == b"PK"  # a real docx is a zip


async def test_export_conversation_as_markdown_and_json(
    client: AsyncClient, conversation_id: str
) -> None:
    await client.post(
        "/api/chat/stream", json={"conversation_id": conversation_id, "content": "hello"}
    )
    md = await client.get(f"/api/exports/conversation/{conversation_id}", params={"format": "md"})
    assert md.status_code == 200
    assert "# " in md.text

    js = await client.get(f"/api/exports/conversation/{conversation_id}", params={"format": "json"})
    assert js.status_code == 200
    assert js.json()["schema"] == "kimi.conversation/1"


async def test_export_unknown_conversation_is_404(client: AsyncClient) -> None:
    resp = await client.get("/api/exports/conversation/nope")
    assert resp.status_code == 404


async def test_arabic_title_survives_the_content_disposition_header(
    client: AsyncClient,
) -> None:
    convo = (await client.post("/api/conversations", json={"title": "تقرير ليبيا"})).json()
    resp = await client.get(f"/api/exports/conversation/{convo['id']}", params={"format": "md"})
    assert resp.status_code == 200
    # RFC 5987 encoding, not a mangled ASCII fallback.
    assert "filename*=UTF-8''" in resp.headers["content-disposition"]
